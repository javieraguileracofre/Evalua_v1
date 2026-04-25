# services/leasing_financiero_export.py
# -*- coding: utf-8 -*-
from __future__ import annotations

import io
from decimal import Decimal
from typing import TYPE_CHECKING, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schemas.comercial.leasing_amortizacion import AmortizacionCuota

if TYPE_CHECKING:
    from models.comercial.leasing_financiero_cotizacion import LeasingFinancieroCotizacion


def build_amortizacion_excel(
    cotizacion: "LeasingFinancieroCotizacion",
    tabla: List[AmortizacionCuota],
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amortización"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).value = f"Tabla de Amortización - Cotización #{cotizacion.id}"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).alignment = center
    row += 2

    ws.cell(row=row, column=1, value="Cliente")
    ws.cell(row=row, column=2, value=(cotizacion.cliente.razon_social if cotizacion.cliente else "-"))
    ws.cell(row=row + 1, column=1, value="Monto financiado")
    ws.cell(
        row=row + 1,
        column=2,
        value=float(cotizacion.monto_financiado or cotizacion.valor_neto or cotizacion.monto or 0),
    )
    ws.cell(row=row + 2, column=1, value="Moneda")
    ws.cell(row=row + 2, column=2, value=cotizacion.moneda)
    ws.cell(row=row + 3, column=1, value="Tasa anual")
    ws.cell(
        row=row + 3,
        column=2,
        value=float((cotizacion.tasa or Decimal("0")) * Decimal("100")),
    )
    ws.cell(row=row + 3, column=3, value="%")
    ws.cell(row=row + 4, column=1, value="Plazo")
    ws.cell(row=row + 4, column=2, value=cotizacion.plazo or 0)
    ws.cell(row=row + 4, column=3, value="periodos")
    ws.cell(row=row + 5, column=1, value="Opción compra")
    ws.cell(row=row + 5, column=2, value=float(cotizacion.opcion_compra or 0))

    for r in range(row, row + 6):
        ws.cell(row=r, column=1).font = bold

    row += 7

    headers = [
        "#",
        "Fecha",
        "Saldo Inicial",
        "Cuota",
        "Interés",
        "Amortización",
        "Saldo Final",
        "Tipo",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = bold
        cell.alignment = center
        cell.border = thin_border
        cell.fill = header_fill
    row += 1

    for cuota in tabla:
        ws.cell(row=row, column=1, value=cuota.numero_cuota).alignment = center
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=str(cuota.fecha_cuota) if cuota.fecha_cuota else "-").alignment = center
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=float(cuota.saldo_inicial)).alignment = right
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4, value=float(cuota.cuota)).alignment = right
        ws.cell(row=row, column=4).border = thin_border

        ws.cell(row=row, column=5, value=float(cuota.interes)).alignment = right
        ws.cell(row=row, column=5).border = thin_border

        ws.cell(row=row, column=6, value=float(cuota.amortizacion)).alignment = right
        ws.cell(row=row, column=6).border = thin_border

        ws.cell(row=row, column=7, value=float(cuota.saldo_final)).alignment = right
        ws.cell(row=row, column=7).border = thin_border

        if cuota.es_gracia:
            tipo = "Gracia"
        elif cuota.es_opcion_compra:
            tipo = "Opción compra"
        else:
            tipo = "Cuota"

        ws.cell(row=row, column=8, value=tipo).alignment = center
        ws.cell(row=row, column=8).border = thin_border

        row += 1

    widths = {1: 5, 2: 12, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16, 8: 14}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_amortizacion_pdf(
    cotizacion: "LeasingFinancieroCotizacion",
    tabla: List[AmortizacionCuota],
) -> io.BytesIO:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Exportación PDF requiere la dependencia 'reportlab'. "
            "Ejecute: pip install reportlab"
        ) from exc

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()
    story = []

    titulo = f"Tabla de Amortización - Cotización #{cotizacion.id}"
    story.append(Paragraph(titulo, styles["Title"]))
    story.append(Spacer(1, 12))

    cliente = cotizacion.cliente.razon_social if cotizacion.cliente else "-"
    monto_fin = cotizacion.monto_financiado or cotizacion.valor_neto or cotizacion.monto or 0
    tasa = cotizacion.tasa or Decimal("0")
    plazo = cotizacion.plazo or 0
    opcion = cotizacion.opcion_compra or 0

    datos_html = (
        f"<b>Cliente:</b> {cliente}<br/>"
        f"<b>Monto financiado:</b> {monto_fin:,.0f} {cotizacion.moneda}<br/>"
        f"<b>Tasa anual:</b> {float(tasa * 100):.2f} %<br/>"
        f"<b>Plazo:</b> {plazo} periodos<br/>"
        f"<b>Opción compra:</b> {opcion:,.0f} {cotizacion.moneda}<br/>"
    )
    story.append(Paragraph(datos_html, styles["Normal"]))
    story.append(Spacer(1, 12))

    data = [
        [
            "#",
            "Fecha",
            "Saldo Inicial",
            "Cuota",
            "Interés",
            "Amortización",
            "Saldo Final",
            "Tipo",
        ]
    ]

    for c in tabla:
        if c.es_gracia:
            tipo = "Gracia"
        elif c.es_opcion_compra:
            tipo = "Opción compra"
        else:
            tipo = "Cuota"

        data.append(
            [
                c.numero_cuota,
                str(c.fecha_cuota) if c.fecha_cuota else "-",
                f"{c.saldo_inicial:,.0f}",
                f"{c.cuota:,.0f}",
                f"{c.interes:,.0f}",
                f"{c.amortizacion:,.0f}",
                f"{c.saldo_final:,.0f}",
                tipo,
            ]
        )

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("ALIGN", (0, 1), (0, -1), "CENTER"),
                ("ALIGN", (2, 1), (6, -1), "RIGHT"),
                ("ALIGN", (1, 1), (1, -1), "CENTER"),
                ("ALIGN", (7, 1), (7, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("TOPPADDING", (0, 0), (-1, 0), 6),
            ]
        )
    )

    story.append(table)

    doc.build(story)
    buffer.seek(0)
    return buffer
