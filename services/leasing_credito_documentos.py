# services/leasing_credito_documentos.py
# -*- coding: utf-8 -*-
"""Carga y extracción de documentos del cliente para scoring leasing."""
from __future__ import annotations

import csv
import hashlib
import io
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from core.paths import PROJECT_ROOT

TIPOS_DOCUMENTO = (
    "CARPETA_TRIBUTARIA",
    "CERTIFICADO_IVA",
    "BALANCE_GENERAL",
    "OTRO",
)

ETIQUETAS_DOCUMENTO = {
    "CARPETA_TRIBUTARIA": "Carpeta tributaria (12 meses)",
    "CERTIFICADO_IVA": "Certificado / F29 IVA",
    "BALANCE_GENERAL": "Último balance general",
    "OTRO": "Otro documento",
}

ALLOWED_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".csv", ".txt", ".png", ".jpg", ".jpeg", ".webp"}
MAX_FILE_BYTES = 15 * 1024 * 1024
IVA_TASA = Decimal("0.19")

STORAGE_ROOT = PROJECT_ROOT / "storage" / "leasing_credito"


def _q(v: Decimal | float | int | str | None, default: str = "0") -> Decimal:
    if v is None or v == "":
        return Decimal(default)
    if isinstance(v, Decimal):
        return v.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if isinstance(v, (int, float)):
        return Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    s = str(v).strip().replace(" ", "").replace("$", "")
    try:
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "," in s and "." not in s:
            parts = s.split(",")
            if len(parts) == 2 and len(parts[1]) <= 2:
                s = s.replace(",", ".")
            else:
                s = s.replace(",", "")
        return Decimal(s).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _norm_header(h: Any) -> str:
    s = str(h or "").strip().lower()
    s = (
        s.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
    )
    return re.sub(r"[^a-z0-9]+", "_", s).strip("_")


def _cell_number(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        return _q(v)
    s = str(v).strip()
    if not s or s in {"-", "—", "N/A", "n/a"}:
        return None
    s = s.replace("$", "").replace("CLP", "").strip()
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s and "." not in s:
        # 1234,56 o 1.234.567 ambiguo: si hay 1 coma y 2 decimales → decimal
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s or s in {".", "-", "-."}:
        return None
    try:
        return _q(s)
    except Exception:
        return None


@dataclass
class ExtraccionDocumento:
    campos: dict[str, Decimal] = field(default_factory=dict)
    alertas: list[str] = field(default_factory=list)
    detalle: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "campos": {k: float(v) for k, v in self.campos.items()},
            "alertas": self.alertas,
            "detalle": self.detalle,
        }


def storage_dir_for(cotizacion_id: int) -> Path:
    path = STORAGE_ROOT / str(int(cotizacion_id))
    path.mkdir(parents=True, exist_ok=True)
    return path


def save_upload_bytes(
    *,
    cotizacion_id: int,
    tipo_documento: str,
    filename: str,
    content: bytes,
) -> tuple[Path, str, str]:
    """Guarda archivo y retorna (path absoluto, path relativo, sha256)."""
    if len(content) > MAX_FILE_BYTES:
        raise ValueError(f"Archivo supera el máximo permitido ({MAX_FILE_BYTES // (1024 * 1024)} MB).")
    ext = Path(filename or "documento.bin").suffix.lower()
    if ext and ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f"Extensión no permitida: {ext}. Use PDF, Excel, CSV o imagen.")
    tipo = (tipo_documento or "OTRO").strip().upper()
    if tipo not in TIPOS_DOCUMENTO:
        tipo = "OTRO"
    digest = hashlib.sha256(content).hexdigest()
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(filename or "documento").name)[:180]
    target_name = f"{tipo}_{digest[:12]}_{safe_name}"
    dest = storage_dir_for(cotizacion_id) / target_name
    dest.write_bytes(content)
    rel = str(dest.relative_to(PROJECT_ROOT)).replace("\\", "/")
    return dest, rel, digest


def parse_documento(content: bytes, filename: str, tipo_documento: str) -> ExtraccionDocumento:
    ext = Path(filename or "").suffix.lower()
    tipo = (tipo_documento or "").strip().upper()
    if ext == ".csv":
        return _parse_tabular_rows(_read_csv_rows(content), tipo)
    if ext in {".xlsx", ".xls"}:
        return _parse_tabular_rows(_read_xlsx_rows(content), tipo)
    if ext == ".txt":
        return _parse_tabular_rows(_read_csv_rows(content), tipo)
    out = ExtraccionDocumento()
    out.alertas.append(
        "Archivo almacenado. PDF/imagen no se parsean automáticamente: complete los montos del documento en el formulario."
    )
    out.detalle = {"parseado": False, "extension": ext, "tipo": tipo}
    return out


def _read_csv_rows(content: bytes) -> list[list[Any]]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = content.decode("utf-8", errors="ignore")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), dialect)
    return [list(row) for row in reader]


def _read_xlsx_rows(content: bytes) -> list[list[Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
    rows: list[list[Any]] = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                if row is None:
                    continue
                vals = list(row)
                if any(v is not None and str(v).strip() != "" for v in vals):
                    rows.append(vals)
                if len(rows) >= 5000:
                    break
            if len(rows) >= 5000:
                break
    finally:
        wb.close()
    return rows


def _parse_tabular_rows(rows: list[list[Any]], tipo: str) -> ExtraccionDocumento:
    out = ExtraccionDocumento(detalle={"parseado": True, "filas": len(rows), "tipo": tipo})
    if not rows:
        out.alertas.append("Archivo tabular vacío.")
        return out

    # Detectar fila de encabezados
    header_idx = 0
    headers_norm: list[str] = []
    for i, row in enumerate(rows[:30]):
        norms = [_norm_header(c) for c in row]
        score = sum(
            1
            for n in norms
            if any(
                k in n
                for k in (
                    "venta",
                    "iva",
                    "debito",
                    "credito",
                    "periodo",
                    "mes",
                    "activo",
                    "pasivo",
                    "patrimonio",
                    "deuda",
                    "ebitda",
                    "utilidad",
                    "monto",
                    "neto",
                )
            )
        )
        if score >= 2:
            header_idx = i
            headers_norm = norms
            break
    if not headers_norm:
        # Formato etiqueta|valor (balance típico)
        return _parse_label_value_rows(rows, tipo)

    data_rows = rows[header_idx + 1 :]
    colmap = {h: idx for idx, h in enumerate(headers_norm) if h}

    def col(*keys: str) -> int | None:
        for k, idx in colmap.items():
            if any(key in k for key in keys):
                return idx
        return None

    if tipo in {"CARPETA_TRIBUTARIA", "CERTIFICADO_IVA", "OTRO"}:
        idx_ventas = col("venta", "ingresos", "neto")
        idx_debito = col("debito", "iva_debito", "iva_deb")
        idx_credito = col("credito", "iva_credito", "iva_cred")
        idx_monto = col("monto", "total", "valor")

        sum_ventas = Decimal("0")
        sum_debito = Decimal("0")
        sum_credito = Decimal("0")
        n = 0
        for row in data_rows:
            if not row:
                continue
            v = _cell_number(row[idx_ventas]) if idx_ventas is not None and idx_ventas < len(row) else None
            d = _cell_number(row[idx_debito]) if idx_debito is not None and idx_debito < len(row) else None
            c = _cell_number(row[idx_credito]) if idx_credito is not None and idx_credito < len(row) else None
            m = _cell_number(row[idx_monto]) if idx_monto is not None and idx_monto < len(row) else None
            if v is None and d is None and c is None and m is None:
                continue
            n += 1
            if v is not None:
                sum_ventas += v
            if d is not None:
                sum_debito += d
            if c is not None:
                sum_credito += c
            if v is None and m is not None and idx_ventas is None:
                sum_ventas += m

        if sum_debito > 0 and sum_ventas <= 0:
            # Proxy ventas afectas desde IVA débito
            sum_ventas = _q(sum_debito / IVA_TASA)
            out.alertas.append("Ventas 12m estimadas desde IVA débito (tasa 19%).")

        if sum_ventas > 0:
            out.campos["ventas_12m_iva"] = sum_ventas
            out.campos["ventas_anuales"] = sum_ventas
        if sum_debito > 0:
            out.campos["iva_debito_12m"] = sum_debito
        if sum_credito > 0:
            out.campos["iva_credito_12m"] = sum_credito
        out.detalle["periodos_detectados"] = n
        if n == 0:
            out.alertas.append("No se detectaron filas numéricas de IVA/ventas; complete manualmente.")

    if tipo in {"BALANCE_GENERAL", "OTRO"}:
        # También intentar label-value en mismas filas por si el header no mapeó bien
        lv = _parse_label_value_rows(rows, "BALANCE_GENERAL")
        for k, v in lv.campos.items():
            out.campos.setdefault(k, v)
        out.alertas.extend(lv.alertas)

        # Columnas explícitas
        for field_name, keys in (
            ("activo_corriente", ("activo_corriente", "activos_corrientes")),
            ("pasivo_corriente", ("pasivo_corriente", "pasivos_corrientes")),
            ("activo_total", ("activo_total", "total_activo", "activos_totales")),
            ("pasivo_total", ("pasivo_total", "total_pasivo", "pasivos_totales")),
            ("patrimonio", ("patrimonio", "capital", "patrimonio_total")),
            ("deuda_financiera_total", ("deuda_financiera", "pasivo_financiero", "obligaciones_financieras")),
            ("utilidad_neta_anual", ("utilidad_neta", "resultado_ejercicio", "ganancia")),
            ("ebitda_anual", ("ebitda",)),
            ("gastos_financieros_anual", ("gastos_financieros", "intereses")),
        ):
            idx = col(*keys)
            if idx is None:
                continue
            for row in data_rows:
                if idx < len(row):
                    num = _cell_number(row[idx])
                    if num is not None and num != 0:
                        out.campos[field_name] = num
                        break

    if not out.campos:
        out.alertas.append("Sin montos extraídos automáticamente; use el formulario de datos del cliente.")
    return out


def _parse_label_value_rows(rows: list[list[Any]], tipo: str) -> ExtraccionDocumento:
    out = ExtraccionDocumento(detalle={"parseado": True, "modo": "etiqueta_valor", "tipo": tipo})
    mapping = (
        ("activo_corriente", ("activo corriente", "activos corrientes")),
        ("pasivo_corriente", ("pasivo corriente", "pasivos corrientes")),
        ("activo_total", ("activo total", "total activo", "total de activos", "activos totales")),
        ("pasivo_total", ("pasivo total", "total pasivo", "total de pasivos", "pasivos totales")),
        ("patrimonio", ("patrimonio", "patrimonio total", "capital propio", "equity")),
        ("deuda_financiera_total", ("deuda financiera", "pasivo financiero", "obligaciones financieras", "prestamos")),
        ("utilidad_neta_anual", ("utilidad neta", "resultado del ejercicio", "ganancia neta")),
        ("ebitda_anual", ("ebitda", "resultado operacional")),
        ("gastos_financieros_anual", ("gastos financieros", "intereses financieros", "costo financiero")),
        ("ventas_anuales", ("ventas", "ingresos", "ingresos operacionales", "ventas netas")),
        ("iva_debito_12m", ("iva debito", "debito fiscal")),
        ("iva_credito_12m", ("iva credito", "credito fiscal")),
    )
    for row in rows:
        if not row or len(row) < 2:
            continue
        label = _norm_header(row[0]).replace("_", " ")
        # buscar primer número en la fila
        num = None
        for cell in row[1:]:
            num = _cell_number(cell)
            if num is not None:
                break
        if num is None:
            continue
        for field_name, aliases in mapping:
            if any(a in label for a in aliases):
                out.campos[field_name] = num
                break
    if "ventas_anuales" in out.campos and "ventas_12m_iva" not in out.campos:
        out.campos["ventas_12m_iva"] = out.campos["ventas_anuales"]
    if "iva_debito_12m" in out.campos and "ventas_anuales" not in out.campos:
        out.campos["ventas_anuales"] = _q(out.campos["iva_debito_12m"] / IVA_TASA)
        out.campos["ventas_12m_iva"] = out.campos["ventas_anuales"]
        out.alertas.append("Ventas estimadas desde IVA débito (19%).")
    return out


def merge_extracciones(docs_datos: list[dict[str, Any]]) -> dict[str, Decimal]:
    """Último valor no nulo por campo gana (docs ordenados asc → el más reciente manda)."""
    merged: dict[str, Decimal] = {}
    for datos in docs_datos:
        campos = (datos or {}).get("campos") or {}
        for k, v in campos.items():
            d = _q(v)
            if d != 0 or k not in merged:
                merged[k] = d
    return merged
