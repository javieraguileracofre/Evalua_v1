# routes/ui/contabilidad.py
# -*- coding: utf-8 -*-
from __future__ import annotations

import calendar
import json
from datetime import date, datetime
from io import BytesIO
from decimal import Decimal
from urllib.parse import urlencode

from fastapi import APIRouter, Depends, Form, Query, Request, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from core.bulk_limits import LIBRO_MAYOR_CONSOLIDADO_MAX_CUENTAS
from core.paths import TEMPLATES_DIR
from core.public_errors import public_error_message
from core.rbac import guard_finanzas_consulta, guard_finanzas_mutacion
from crud.finanzas import contabilidad_asientos as crud_asientos
from crud.finanzas import plan_cuentas as crud_plan
from db.session import get_db
from schemas.finanzas.plan_cuentas import PlanCuentaCreate, PlanCuentaUpdate

router = APIRouter(tags=["Contabilidad"])
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))


def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _autosize_columns(ws) -> None:
    for col in ws.columns:
        if not col:
            continue
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 54)


def _parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def _shift_one_year(d: date) -> date:
    try:
        return d.replace(year=d.year - 1)
    except ValueError:
        # Ajuste de 29-feb a 28-feb al restar año.
        return d.replace(month=2, day=28, year=d.year - 1)


def _resolver_periodos_comparables(
    fecha_desde: str | None,
    fecha_hasta: str | None,
) -> dict[str, str]:
    hoy = date.today()
    desde_in = _parse_iso_date(fecha_desde)
    hasta_in = _parse_iso_date(fecha_hasta)

    if desde_in and hasta_in:
        actual_desde = desde_in
        actual_hasta = hasta_in
    elif hasta_in:
        actual_desde = date(hasta_in.year, 1, 1)
        actual_hasta = hasta_in
    elif desde_in:
        actual_desde = desde_in
        actual_hasta = hoy
    else:
        actual_desde = date(hoy.year, 1, 1)
        actual_hasta = hoy

    previo_desde = _shift_one_year(actual_desde)
    previo_hasta = _shift_one_year(actual_hasta)

    return {
        "actual_desde": actual_desde.isoformat(),
        "actual_hasta": actual_hasta.isoformat(),
        "previo_desde": previo_desde.isoformat(),
        "previo_hasta": previo_hasta.isoformat(),
        "anio_actual": str(actual_hasta.year),
        "anio_previo": str(previo_hasta.year),
    }


def _rows_comparables(actual_rows: list[dict], previo_rows: list[dict]) -> list[dict]:
    def _key(item: dict) -> tuple[str, str]:
        return (
            str(item.get("codigo_cuenta") or "").strip(),
            str(item.get("nombre_cuenta") or "").strip(),
        )

    actual_map = {_key(r): r for r in actual_rows}
    previo_map = {_key(r): r for r in previo_rows}
    keys = sorted(set(actual_map.keys()) | set(previo_map.keys()))

    out: list[dict] = []
    for key in keys:
        ar = actual_map.get(key, {})
        pr = previo_map.get(key, {})
        monto_actual = Decimal(str(ar.get("monto") or 0))
        monto_previo = Decimal(str(pr.get("monto") or 0))
        variacion = monto_actual - monto_previo
        variacion_pct = None if monto_previo == 0 else (variacion / monto_previo)
        out.append(
            {
                "codigo_cuenta": key[0],
                "nombre_cuenta": key[1],
                "monto_actual": monto_actual,
                "monto_previo": monto_previo,
                "variacion_pct": variacion_pct,
            }
        )
    return out


def _format_balance_8_columnas_excel(
    ws,
    *,
    data: dict,
    fecha_desde: str,
    fecha_hasta: str,
) -> None:
    title_font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    head_font = Font(name="Calibri", size=10, bold=True, color="1F2937")
    bold_font = Font(name="Calibri", size=10, bold=True, color="111827")
    normal_font = Font(name="Calibri", size=10, color="111827")
    fill_title = PatternFill(fill_type="solid", fgColor="0F4C81")
    fill_head = PatternFill(fill_type="solid", fgColor="DDEBF7")
    fill_total = PatternFill(fill_type="solid", fgColor="EDF4FC")
    thin = Side(style="thin", color="D1D5DB")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.title = "Balance8Columnas"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    ws.merge_cells("A1:H1")
    ws["A1"] = "BALANCE DE 8 COLUMNAS"
    ws["A1"].font = title_font
    ws["A1"].fill = fill_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Período: {fecha_desde} a {fecha_hasta}"
    ws["A2"].font = Font(name="Calibri", size=10, color="FFFFFF", italic=True)
    ws["A2"].fill = fill_title
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "Código",
        "Cuenta",
        "Saldo inicial deudor",
        "Saldo inicial acreedor",
        "Movimiento debe",
        "Movimiento haber",
        "Saldo final deudor",
        "Saldo final acreedor",
    ]
    row = 4
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=h)
        c.font = head_font
        c.fill = fill_head
        c.border = border_all
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    for item in data.get("rows", []):
        ws.cell(row=row, column=1, value=item.get("codigo_cuenta") or "").font = normal_font
        ws.cell(row=row, column=2, value=item.get("nombre_cuenta") or "").font = normal_font
        ws.cell(row=row, column=3, value=float(item.get("saldo_inicial_deudor") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
        ws.cell(row=row, column=4, value=float(item.get("saldo_inicial_acreedor") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
        ws.cell(row=row, column=5, value=float(item.get("movimiento_debe") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
        ws.cell(row=row, column=6, value=float(item.get("movimiento_haber") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
        ws.cell(row=row, column=7, value=float(item.get("saldo_final_deudor") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
        ws.cell(row=row, column=8, value=float(item.get("saldo_final_acreedor") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
        for col in range(1, 9):
            c = ws.cell(row=row, column=col)
            c.border = border_all
            if col >= 3:
                c.alignment = Alignment(horizontal="right", vertical="center")
        row += 1

    tot = data.get("totales", {})
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="Totales").font = bold_font
    ws.cell(row=row, column=3, value=float(tot.get("saldo_inicial_deudor") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
    ws.cell(row=row, column=4, value=float(tot.get("saldo_inicial_acreedor") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
    ws.cell(row=row, column=5, value=float(tot.get("movimiento_debe") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
    ws.cell(row=row, column=6, value=float(tot.get("movimiento_haber") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
    ws.cell(row=row, column=7, value=float(tot.get("saldo_final_deudor") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
    ws.cell(row=row, column=8, value=float(tot.get("saldo_final_acreedor") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
    for col in range(1, 9):
        c = ws.cell(row=row, column=col)
        c.font = bold_font
        c.fill = fill_total
        c.border = border_all
        if col >= 3:
            c.alignment = Alignment(horizontal="right", vertical="center")

    _autosize_columns(ws)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 14)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 42)


def _format_libro_mayor_excel(ws, *, data: dict) -> None:
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    head_font = Font(name="Calibri", size=10, bold=True, color="1F2937")
    bold_font = Font(name="Calibri", size=10, bold=True, color="111827")
    normal_font = Font(name="Calibri", size=10, color="111827")
    fill_title = PatternFill(fill_type="solid", fgColor="0F4C81")
    fill_head = PatternFill(fill_type="solid", fgColor="DDEBF7")
    fill_total = PatternFill(fill_type="solid", fgColor="EDF4FC")
    thin = Side(style="thin", color="D1D5DB")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.title = "LibroMayor"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    cuenta = f"{data.get('codigo_cuenta') or ''} - {data.get('nombre_cuenta') or ''}".strip(" -")
    ws.merge_cells("A1:F1")
    ws["A1"] = "LIBRO MAYOR POR CUENTA"
    ws["A1"].font = title_font
    ws["A1"].fill = fill_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 23

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Cuenta: {cuenta}"
    ws["A2"].font = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
    ws["A2"].fill = fill_title
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:F3")
    ws["A3"] = (
        f"Período: {data.get('fecha_desde') or 'inicio'} a {data.get('fecha_hasta') or 'hoy'} "
        f"| Saldo inicial: {float(data.get('saldo_inicial') or 0):,.2f}"
    )
    ws["A3"].font = Font(name="Calibri", size=10, color="FFFFFF", italic=True)
    ws["A3"].fill = fill_title
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Fecha", "Asiento", "Glosa", "Detalle", "Debe", "Haber", "Saldo"]
    row = 5
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=h)
        c.font = head_font
        c.fill = fill_head
        c.border = border_all
        c.alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for item in data.get("rows", []):
        ws.cell(row=row, column=1, value=str(item.get("fecha") or "")[:10]).font = normal_font
        ws.cell(row=row, column=2, value=item.get("asiento_id")).font = normal_font
        ws.cell(row=row, column=3, value=item.get("glosa") or "").font = normal_font
        ws.cell(row=row, column=4, value=item.get("detalle") or "").font = normal_font
        ws.cell(row=row, column=5, value=float(item.get("debe") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
        ws.cell(row=row, column=6, value=float(item.get("haber") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
        ws.cell(row=row, column=7, value=float(item.get("saldo") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
        for col in range(1, 8):
            c = ws.cell(row=row, column=col)
            c.border = border_all
            if col >= 5:
                c.alignment = Alignment(horizontal="right", vertical="center")
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="Totales período").font = bold_font
    ws.cell(row=row, column=5, value=float(data.get("total_debe") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
    ws.cell(row=row, column=6, value=float(data.get("total_haber") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
    ws.cell(row=row, column=7, value=float(data.get("saldo_final") or 0)).number_format = '#,##0.00;[Red]-#,##0.00'
    for col in range(1, 8):
        c = ws.cell(row=row, column=col)
        c.font = bold_font
        c.fill = fill_total
        c.border = border_all
        if col >= 5:
            c.alignment = Alignment(horizontal="right", vertical="center")

    _autosize_columns(ws)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width or 0, 34)
    ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width or 0, 30)


def _safe_sheet_title(raw: str, *, fallback: str = "Hoja") -> str:
    banned = set(r'[]:*?/\\')
    clean = "".join("_" if ch in banned else ch for ch in (raw or "")).strip()
    if not clean:
        clean = fallback
    return clean[:31]


def _format_estado_resultados_excel(
    ws,
    *,
    resumen_actual: dict,
    resumen_previo: dict,
    periodo_actual: str,
    anio_actual: str,
    anio_previo: str,
) -> None:
    # Estilo corporativo simple y legible para impresión o análisis en Excel.
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    head_font = Font(name="Calibri", size=10, bold=True, color="1F2937")
    bold_font = Font(name="Calibri", size=10, bold=True, color="111827")
    normal_font = Font(name="Calibri", size=10, color="111827")

    fill_title = PatternFill(fill_type="solid", fgColor="0F4C81")
    fill_section = PatternFill(fill_type="solid", fgColor="1F6AA5")
    fill_head = PatternFill(fill_type="solid", fgColor="E7EEF7")
    fill_total = PatternFill(fill_type="solid", fgColor="F4F8FC")
    fill_result = PatternFill(fill_type="solid", fgColor="DDEBF7")
    thin = Side(style="thin", color="D1D5DB")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.title = "EstadoResultados"
    ws.freeze_panes = "A8"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "ESTADO DE RESULTADOS"
    ws["A1"].font = title_font
    ws["A1"].fill = fill_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    periodo_label = f"Período: {periodo_actual}"
    ws.merge_cells("A2:E2")
    ws["A2"] = periodo_label
    ws["A2"].font = subtitle_font
    ws["A2"].fill = fill_title
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:E3")
    ws["A3"] = f"Generado: {date.today().isoformat()}"
    ws["A3"].font = Font(name="Calibri", size=9, color="FFFFFF")
    ws["A3"].fill = fill_title
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 16

    row = 5

    def write_block(
        title: str,
        rows_actual: list[dict],
        rows_previo: list[dict],
        total_actual: Decimal,
        total_previo: Decimal,
    ) -> int:
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=title)
        c.font = section_font
        c.fill = fill_section
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border_all
        row += 1

        ws.cell(row=row, column=1, value="Código").font = head_font
        ws.cell(row=row, column=2, value="Cuenta").font = head_font
        ws.cell(row=row, column=3, value=anio_actual).font = head_font
        ws.cell(row=row, column=4, value=anio_previo).font = head_font
        ws.cell(row=row, column=5, value="Var%").font = head_font
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill_head
            cell.border = border_all
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")
        row += 1

        comparables = _rows_comparables(rows_actual, rows_previo)
        if comparables:
            for item in comparables:
                ws.cell(row=row, column=1, value=item.get("codigo_cuenta") or "").font = normal_font
                ws.cell(row=row, column=2, value=item.get("nombre_cuenta") or "").font = normal_font
                c_actual = ws.cell(row=row, column=3, value=float(item.get("monto_actual") or 0))
                c_actual.font = normal_font
                c_actual.number_format = '#,##0.00;[Red]-#,##0.00'
                c_actual.alignment = Alignment(horizontal="right", vertical="center")
                c_previo = ws.cell(row=row, column=4, value=float(item.get("monto_previo") or 0))
                c_previo.font = normal_font
                c_previo.number_format = '#,##0.00;[Red]-#,##0.00'
                c_previo.alignment = Alignment(horizontal="right", vertical="center")

                pct = item.get("variacion_pct")
                if pct is None:
                    c_var = ws.cell(row=row, column=5, value="-")
                    c_var.font = Font(name="Calibri", size=10, color="6B7280", italic=True)
                    c_var.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c_var = ws.cell(row=row, column=5, value=float(pct))
                    c_var.font = normal_font
                    c_var.number_format = "0.00%"
                    c_var.alignment = Alignment(horizontal="right", vertical="center")

                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = border_all
                row += 1
        else:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            empty = ws.cell(row=row, column=1, value="Sin movimientos en el período")
            empty.font = Font(name="Calibri", size=10, italic=True, color="6B7280")
            empty.alignment = Alignment(horizontal="center", vertical="center")
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border_all
            row += 1

        ws.cell(row=row, column=1, value=f"Total {title.lower()}").font = bold_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        t_actual = ws.cell(row=row, column=3, value=float(total_actual or 0))
        t_actual.font = bold_font
        t_actual.number_format = '#,##0.00;[Red]-#,##0.00'
        t_actual.alignment = Alignment(horizontal="right", vertical="center")
        t_previo = ws.cell(row=row, column=4, value=float(total_previo or 0))
        t_previo.font = bold_font
        t_previo.number_format = '#,##0.00;[Red]-#,##0.00'
        t_previo.alignment = Alignment(horizontal="right", vertical="center")
        if total_previo == 0:
            t_var = ws.cell(row=row, column=5, value="-")
            t_var.font = Font(name="Calibri", size=10, color="6B7280", italic=True)
        else:
            t_var = ws.cell(row=row, column=5, value=float((total_actual - total_previo) / total_previo))
            t_var.font = bold_font
            t_var.number_format = "0.00%"
        t_var.alignment = Alignment(horizontal="right", vertical="center")

        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = fill_total
            ws.cell(row=row, column=col).border = border_all
        row += 2
        return row

    write_block(
        "Ingresos",
        resumen_actual.get("ingresos", []),
        resumen_previo.get("ingresos", []),
        Decimal(str(resumen_actual.get("total_ingresos") or 0)),
        Decimal(str(resumen_previo.get("total_ingresos") or 0)),
    )
    write_block(
        "Costos",
        resumen_actual.get("costos", []),
        resumen_previo.get("costos", []),
        Decimal(str(resumen_actual.get("total_costos") or 0)),
        Decimal(str(resumen_previo.get("total_costos") or 0)),
    )
    write_block(
        "Gastos",
        resumen_actual.get("gastos", []),
        resumen_previo.get("gastos", []),
        Decimal(str(resumen_actual.get("total_gastos") or 0)),
        Decimal(str(resumen_previo.get("total_gastos") or 0)),
    )

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="Utilidad bruta (Ingresos - Costos)").font = bold_font
    ub_actual_num = Decimal(str(resumen_actual.get("utilidad_bruta") or 0))
    ub_previo_num = Decimal(str(resumen_previo.get("utilidad_bruta") or 0))
    ub = ws.cell(row=row, column=3, value=float(ub_actual_num))
    ub.font = bold_font
    ub.number_format = '#,##0.00;[Red]-#,##0.00'
    ub.alignment = Alignment(horizontal="right", vertical="center")
    ub_prev = ws.cell(row=row, column=4, value=float(ub_previo_num))
    ub_prev.font = bold_font
    ub_prev.number_format = '#,##0.00;[Red]-#,##0.00'
    ub_prev.alignment = Alignment(horizontal="right", vertical="center")
    if ub_previo_num == 0:
        ub_var = ws.cell(row=row, column=5, value="-")
        ub_var.font = Font(name="Calibri", size=10, color="6B7280", italic=True)
    else:
        ub_var = ws.cell(row=row, column=5, value=float((ub_actual_num - ub_previo_num) / ub_previo_num))
        ub_var.font = bold_font
        ub_var.number_format = "0.00%"
    ub_var.alignment = Alignment(horizontal="right", vertical="center")
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = fill_result
        ws.cell(row=row, column=col).border = border_all
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="Resultado operacional").font = Font(
        name="Calibri", size=11, bold=True, color="0F172A"
    )
    ro_actual_num = Decimal(str(resumen_actual.get("resultado_operacional") or 0))
    ro_previo_num = Decimal(str(resumen_previo.get("resultado_operacional") or 0))
    ro = ws.cell(row=row, column=3, value=float(ro_actual_num))
    ro.font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    ro.number_format = '#,##0.00;[Red]-#,##0.00'
    ro.alignment = Alignment(horizontal="right", vertical="center")
    ro_prev = ws.cell(row=row, column=4, value=float(ro_previo_num))
    ro_prev.font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    ro_prev.number_format = '#,##0.00;[Red]-#,##0.00'
    ro_prev.alignment = Alignment(horizontal="right", vertical="center")
    if ro_previo_num == 0:
        ro_var = ws.cell(row=row, column=5, value="-")
        ro_var.font = Font(name="Calibri", size=10, color="6B7280", italic=True)
    else:
        ro_var = ws.cell(row=row, column=5, value=float((ro_actual_num - ro_previo_num) / ro_previo_num))
        ro_var.font = Font(name="Calibri", size=11, bold=True, color="0F172A")
        ro_var.number_format = "0.00%"
    ro_var.alignment = Alignment(horizontal="right", vertical="center")
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = fill_result
        ws.cell(row=row, column=col).border = border_all

    _autosize_columns(ws)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 14)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 42)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width or 0, 18)
    ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width or 0, 18)
    ws.column_dimensions["E"].width = max(ws.column_dimensions["E"].width or 0, 12)


def _redirect(
    request: Request,
    route_name: str,
    *,
    msg: str | None = None,
    sev: str = "info",
    status_code: int = status.HTTP_303_SEE_OTHER,
    **path_params,
) -> RedirectResponse:
    url = str(request.url_for(route_name, **path_params))
    params = {}
    if msg:
        params["msg"] = msg
        params["sev"] = sev

    if params:
        url = f"{url}?{urlencode(params)}"

    return RedirectResponse(url=url, status_code=status_code)


@router.get("/contabilidad/plan-cuentas", response_class=HTMLResponse, name="contabilidad_plan_cuentas")
def contabilidad_plan_cuentas(
    request: Request,
    q: str | None = Query(None),
    tipo: str | None = Query(None),
    estado: str | None = Query(None),
    solo_movimiento: bool | None = Query(None),
    msg: str | None = Query(None),
    sev: str = Query("info"),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    cuentas = crud_plan.listar_plan_cuentas(
        db,
        q=q,
        tipo=tipo,
        estado=estado,
        solo_movimiento=solo_movimiento,
    )
    padres = crud_plan.listar_plan_cuentas(db, estado="ACTIVO")

    return templates.TemplateResponse(
        "finanzas/plan_cuentas_lista.html",
        {
            "request": request,
            "cuentas": cuentas,
            "padres": padres,
            "q": q,
            "tipo": tipo,
            "estado": estado,
            "solo_movimiento": solo_movimiento,
            "msg": msg,
            "sev": sev,
            "active_menu": "contabilidad_plan_cuentas",
        },
    )


@router.get("/contabilidad/plan-cuentas/nuevo", response_class=HTMLResponse, name="contabilidad_plan_cuenta_nuevo")
def contabilidad_plan_cuenta_nuevo(
    request: Request,
    msg: str | None = Query(None),
    sev: str = Query("info"),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_mutacion(request)) is not None:
        return redir
    padres = crud_plan.listar_plan_cuentas(db, estado="ACTIVO")

    return templates.TemplateResponse(
        "finanzas/plan_cuenta_form.html",
        {
            "request": request,
            "modo": "crear",
            "cuenta": None,
            "padres": padres,
            "msg": msg,
            "sev": sev,
            "active_menu": "contabilidad_plan_cuentas",
        },
    )


@router.post("/contabilidad/plan-cuentas/nuevo", name="contabilidad_plan_cuenta_crear")
def contabilidad_plan_cuenta_crear(
    request: Request,
    codigo: str = Form(...),
    nombre: str = Form(...),
    nivel: int = Form(...),
    cuenta_padre_id: int | None = Form(None),
    tipo: str = Form(...),
    clasificacion: str = Form(...),
    naturaleza: str = Form(...),
    acepta_movimiento: str | None = Form(None),
    requiere_centro_costo: str | None = Form(None),
    estado: str = Form(...),
    descripcion: str | None = Form(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_mutacion(request)) is not None:
        return redir
    try:
        payload = PlanCuentaCreate(
            codigo=codigo,
            nombre=nombre,
            nivel=nivel,
            cuenta_padre_id=cuenta_padre_id,
            tipo=tipo,
            clasificacion=clasificacion,
            naturaleza=naturaleza,
            acepta_movimiento=bool(acepta_movimiento),
            requiere_centro_costo=bool(requiere_centro_costo),
            estado=estado,
            descripcion=descripcion,
        )
        crud_plan.crear_plan_cuenta(db, payload)
        return _redirect(
            request,
            "contabilidad_plan_cuentas",
            msg="Cuenta creada correctamente.",
            sev="success",
        )
    except Exception as e:
        return _redirect(
            request,
            "contabilidad_plan_cuenta_nuevo",
            msg=public_error_message(e, default="No se pudo crear la cuenta."),
            sev="danger",
        )


@router.get("/contabilidad/plan-cuentas/{cuenta_id}/editar", response_class=HTMLResponse, name="contabilidad_plan_cuenta_editar")
def contabilidad_plan_cuenta_editar(
    request: Request,
    cuenta_id: int,
    msg: str | None = Query(None),
    sev: str = Query("info"),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_mutacion(request)) is not None:
        return redir
    cuenta = crud_plan.obtener_plan_cuenta(db, cuenta_id)
    if not cuenta:
        return _redirect(
            request,
            "contabilidad_plan_cuentas",
            msg="Cuenta no encontrada.",
            sev="warning",
        )

    padres = crud_plan.listar_plan_cuentas(db, estado="ACTIVO")
    padres = [p for p in padres if p.id != cuenta.id]

    return templates.TemplateResponse(
        "finanzas/plan_cuenta_form.html",
        {
            "request": request,
            "modo": "editar",
            "cuenta": cuenta,
            "padres": padres,
            "msg": msg,
            "sev": sev,
            "active_menu": "contabilidad_plan_cuentas",
        },
    )


@router.post("/contabilidad/plan-cuentas/{cuenta_id}/editar", name="contabilidad_plan_cuenta_actualizar")
def contabilidad_plan_cuenta_actualizar(
    request: Request,
    cuenta_id: int,
    codigo: str = Form(...),
    nombre: str = Form(...),
    nivel: int = Form(...),
    cuenta_padre_id: int | None = Form(None),
    tipo: str = Form(...),
    clasificacion: str = Form(...),
    naturaleza: str = Form(...),
    acepta_movimiento: str | None = Form(None),
    requiere_centro_costo: str | None = Form(None),
    estado: str = Form(...),
    descripcion: str | None = Form(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_mutacion(request)) is not None:
        return redir
    cuenta = crud_plan.obtener_plan_cuenta(db, cuenta_id)
    if not cuenta:
        return _redirect(
            request,
            "contabilidad_plan_cuentas",
            msg="Cuenta no encontrada.",
            sev="warning",
        )

    try:
        payload = PlanCuentaUpdate(
            codigo=codigo,
            nombre=nombre,
            nivel=nivel,
            cuenta_padre_id=cuenta_padre_id,
            tipo=tipo,
            clasificacion=clasificacion,
            naturaleza=naturaleza,
            acepta_movimiento=bool(acepta_movimiento),
            requiere_centro_costo=bool(requiere_centro_costo),
            estado=estado,
            descripcion=descripcion,
        )
        crud_plan.actualizar_plan_cuenta(db, cuenta=cuenta, payload=payload)
        return _redirect(
            request,
            "contabilidad_plan_cuentas",
            msg="Cuenta actualizada correctamente.",
            sev="success",
        )
    except Exception as e:
        return _redirect(
            request,
            "contabilidad_plan_cuenta_editar",
            cuenta_id=cuenta_id,
            msg=public_error_message(e, default="No se pudo actualizar la cuenta."),
            sev="danger",
        )


@router.post("/contabilidad/plan-cuentas/{cuenta_id}/desactivar", name="contabilidad_plan_cuenta_desactivar")
def contabilidad_plan_cuenta_desactivar(
    request: Request,
    cuenta_id: int,
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_mutacion(request)) is not None:
        return redir
    cuenta = crud_plan.obtener_plan_cuenta(db, cuenta_id)
    if not cuenta:
        return _redirect(
            request,
            "contabilidad_plan_cuentas",
            msg="Cuenta no encontrada.",
            sev="warning",
        )

    crud_plan.desactivar_plan_cuenta(db, cuenta=cuenta)
    return _redirect(
        request,
        "contabilidad_plan_cuentas",
        msg="Cuenta desactivada.",
        sev="success",
    )


@router.get("/contabilidad/asientos", response_class=HTMLResponse, name="contabilidad_asientos")
def contabilidad_asientos(
    request: Request,
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    origen_tipo: str | None = Query(None),
    msg: str | None = Query(None),
    sev: str = Query("info"),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    asientos = crud_asientos.listar_asientos(
        db,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        origen_tipo=origen_tipo,
        limit=300,
    )
    return templates.TemplateResponse(
        "finanzas/asientos_lista.html",
        {
            "request": request,
            "asientos": asientos,
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,
            "origen_tipo": origen_tipo,
            "msg": msg,
            "sev": sev,
            "active_menu": "contabilidad_asientos",
        },
    )


@router.get("/contabilidad/asientos/manual/nuevo", response_class=HTMLResponse, name="contabilidad_asiento_manual_nuevo")
def contabilidad_asiento_manual_nuevo(
    request: Request,
    msg: str | None = Query(None),
    sev: str = Query("info"),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_mutacion(request)) is not None:
        return redir
    cuentas = crud_plan.listar_cuentas_movimiento_activas(db)
    return templates.TemplateResponse(
        "finanzas/asiento_manual_form.html",
        {
            "request": request,
            "cuentas": cuentas,
            "msg": msg,
            "sev": sev,
            "active_menu": "contabilidad_asientos",
        },
    )


@router.post("/contabilidad/asientos/manual/nuevo", name="contabilidad_asiento_manual_crear")
def contabilidad_asiento_manual_crear(
    request: Request,
    fecha: date = Form(...),
    glosa: str = Form(...),
    detalles_json: str = Form(...),
    usuario: str | None = Form(None),
    es_apertura: str | None = Form(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_mutacion(request)) is not None:
        return redir
    try:
        lineas = json.loads(detalles_json)
        if not isinstance(lineas, list) or len(lineas) < 2:
            raise ValueError("Debe enviar al menos dos líneas contables.")
        detalles: list[dict] = []
        for ln in lineas:
            if not isinstance(ln, dict):
                continue
            detalles.append(
                {
                    "codigo_cuenta": str(ln.get("codigo_cuenta") or "").strip(),
                    "descripcion": (ln.get("descripcion") or None),
                    "debe": ln.get("debe") or 0,
                    "haber": ln.get("haber") or 0,
                }
            )
        ap_flag = (es_apertura or "").strip().lower() in {"1", "on", "true", "yes"}
        origen_tipo = "MANUAL_APERTURA" if ap_flag else "MANUAL_AJUSTE"
        crud_asientos.crear_asiento(
            db,
            fecha=fecha,
            origen_tipo=origen_tipo,
            origen_id=0,
            glosa=(glosa or "").strip()[:255],
            detalles=detalles,
            usuario=(usuario or "").strip() or None,
        )
    except Exception as exc:
        return _redirect(
            request,
            "contabilidad_asiento_manual_nuevo",
            msg=public_error_message(exc, default="No se pudo registrar el asiento manual."),
            sev="danger",
        )

    return _redirect(
        request,
        "contabilidad_asientos",
        msg="Asiento manual registrado correctamente.",
        sev="success",
    )


@router.get("/contabilidad/asientos/{asiento_id}", response_class=HTMLResponse, name="contabilidad_asiento_detalle")
def contabilidad_asiento_detalle(
    request: Request,
    asiento_id: int,
    msg: str | None = Query(None),
    sev: str = Query("info"),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    data = crud_asientos.obtener_asiento_detalle(db, asiento_id)
    if not data:
        return _redirect(
            request,
            "contabilidad_asientos",
            msg="Asiento no encontrado.",
            sev="warning",
        )

    ot = str(data["cabecera"].get("origen_tipo") or "").upper()
    puede_revertir = ot in crud_asientos.ORIGENES_MANUALES_REVERSIBLES

    return templates.TemplateResponse(
        "finanzas/asiento_detalle.html",
        {
            "request": request,
            "cabecera": data["cabecera"],
            "detalles": data["detalles"],
            "total_debe": data["total_debe"],
            "total_haber": data["total_haber"],
            "puede_revertir": puede_revertir,
            "msg": msg,
            "sev": sev,
            "active_menu": "contabilidad_asientos",
        },
    )


@router.post("/contabilidad/asientos/{asiento_id}/revertir", name="contabilidad_revertir_asiento")
def contabilidad_revertir_asiento(
    request: Request,
    asiento_id: int,
    glosa_reversion: str | None = Form(None),
    usuario: str | None = Form(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_mutacion(request)) is not None:
        return redir
    try:
        nuevo_id = crud_asientos.crear_reversion_asiento_manual(
            db,
            asiento_original_id=asiento_id,
            glosa=glosa_reversion,
            usuario=(usuario or "").strip() or None,
        )
    except Exception as exc:
        return _redirect(
            request,
            "contabilidad_asiento_detalle",
            asiento_id=asiento_id,
            msg=public_error_message(exc, default="No se pudo crear la reversión del asiento."),
            sev="danger",
        )

    return _redirect(
        request,
        "contabilidad_asiento_detalle",
        asiento_id=nuevo_id,
        msg=f"Reversión creada correctamente (asiento #{nuevo_id}).",
        sev="success",
    )


@router.get("/contabilidad/estado-resultados", response_class=HTMLResponse, name="contabilidad_estado_resultados")
def contabilidad_estado_resultados(
    request: Request,
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    resumen = crud_asientos.obtener_estado_resultados(
        db,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    hoy = date.today()
    _, ultimo_dia = calendar.monthrange(hoy.year, hoy.month)
    mes_desde = hoy.replace(day=1).isoformat()
    mes_hasta = hoy.replace(day=ultimo_dia).isoformat()
    ti: Decimal = resumen["total_ingresos"]
    ro: Decimal = resumen["resultado_operacional"]
    # Texto ya formateado: evita str.format sobre Undefined u objetos raros en Jinja.
    pct_ro_sobre_ingresos_pct: str | None
    if ti != 0:
        pct_ro_sobre_ingresos_pct = f"{float(ro / ti * Decimal('100')):.1f}"
    else:
        pct_ro_sobre_ingresos_pct = None

    return templates.TemplateResponse(
        "finanzas/estado_resultados.html",
        {
            "request": request,
            "resumen": resumen,
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,
            "fecha_hoy": hoy.isoformat(),
            "er_mes_desde": mes_desde,
            "er_mes_hasta": mes_hasta,
            "pct_ro_sobre_ingresos_pct": pct_ro_sobre_ingresos_pct,
            "active_menu": "contabilidad_estado_resultados",
        },
    )


@router.get(
    "/contabilidad/estado-resultados/export.xlsx",
    name="contabilidad_estado_resultados_excel",
)
def contabilidad_estado_resultados_excel(
    request: Request,
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    periodos = _resolver_periodos_comparables(fecha_desde, fecha_hasta)

    resumen_actual = crud_asientos.obtener_estado_resultados(
        db,
        fecha_desde=periodos["actual_desde"],
        fecha_hasta=periodos["actual_hasta"],
    )
    resumen_previo = crud_asientos.obtener_estado_resultados(
        db,
        fecha_desde=periodos["previo_desde"],
        fecha_hasta=periodos["previo_hasta"],
    )

    wb = Workbook()
    ws = wb.active
    _format_estado_resultados_excel(
        ws,
        resumen_actual=resumen_actual,
        resumen_previo=resumen_previo,
        periodo_actual=f"{periodos['actual_desde']} a {periodos['actual_hasta']}",
        anio_actual=periodos["anio_actual"],
        anio_previo=periodos["anio_previo"],
    )

    filename = (
        f"estado_resultados_{periodos['anio_actual']}_vs_{periodos['anio_previo']}_"
        f"{periodos['actual_desde']}_{periodos['actual_hasta']}.xlsx"
    )
    return _xlsx_response(wb, filename)


@router.get(
    "/contabilidad/balance-general/export-8-columnas.xlsx",
    name="contabilidad_balance_8_columnas_excel",
)
def contabilidad_balance_8_columnas_excel(
    request: Request,
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    fh = _parse_iso_date(fecha_hasta) or date.today()
    fd = _parse_iso_date(fecha_desde) or date(fh.year, 1, 1)
    if fd > fh:
        fd = date(fh.year, 1, 1)

    data = crud_asientos.obtener_balance_8_columnas(
        db,
        fecha_desde=fd.isoformat(),
        fecha_hasta=fh.isoformat(),
    )
    wb = Workbook()
    ws = wb.active
    _format_balance_8_columnas_excel(
        ws,
        data=data,
        fecha_desde=fd.isoformat(),
        fecha_hasta=fh.isoformat(),
    )
    filename = f"balance_8_columnas_{fd.isoformat()}_{fh.isoformat()}.xlsx"
    return _xlsx_response(wb, filename)


@router.get(
    "/contabilidad/libro-mayor/export.xlsx",
    name="contabilidad_libro_mayor_excel",
)
def contabilidad_libro_mayor_excel(
    request: Request,
    codigo_cuenta: str = Query(...),
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    data = crud_asientos.obtener_libro_mayor_cuenta(
        db,
        codigo_cuenta=codigo_cuenta,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    wb = Workbook()
    ws = wb.active
    _format_libro_mayor_excel(ws, data=data)
    filename = f"libro_mayor_{codigo_cuenta}_{fecha_desde or 'inicio'}_{fecha_hasta or 'hoy'}.xlsx"
    return _xlsx_response(wb, filename)


@router.get(
    "/contabilidad/libro-mayor-consolidado/export.xlsx",
    name="contabilidad_libro_mayor_consolidado_excel",
)
def contabilidad_libro_mayor_consolidado_excel(
    request: Request,
    codigo_desde: str | None = Query(None),
    codigo_hasta: str | None = Query(None),
    tipo: str | None = Query(None),
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    cuentas = crud_plan.listar_cuentas_movimiento_activas(db)
    tipo_filter = (tipo or "").strip().upper()
    c_desde = (codigo_desde or "").strip()
    c_hasta = (codigo_hasta or "").strip()

    seleccionadas = []
    for c in cuentas:
        codigo = str(getattr(c, "codigo", "") or "").strip()
        if not codigo:
            continue
        if tipo_filter and str(getattr(c, "tipo", "") or "").upper() != tipo_filter:
            continue
        if c_desde and codigo < c_desde:
            continue
        if c_hasta and codigo > c_hasta:
            continue
        seleccionadas.append(c)

    total_cuentas_filtro = len(seleccionadas)
    truncado_consolidado = total_cuentas_filtro > LIBRO_MAYOR_CONSOLIDADO_MAX_CUENTAS
    if truncado_consolidado:
        seleccionadas = seleccionadas[:LIBRO_MAYOR_CONSOLIDADO_MAX_CUENTAS]

    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.sheet_view.showGridLines = False
    ws_resumen.append(
        [
            "Código",
            "Cuenta",
            "Tipo",
            "Movimientos",
            "Total Debe",
            "Total Haber",
            "Saldo Final",
        ]
    )
    for c in ws_resumen[1]:
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill(fill_type="solid", fgColor="DDEBF7")

    if truncado_consolidado:
        ws_resumen.append(
            [
                f"(Límite de exportación: {LIBRO_MAYOR_CONSOLIDADO_MAX_CUENTAS} de {total_cuentas_filtro} cuentas; "
                "refine tipo o rango de códigos y vuelva a exportar.)",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )

    if not seleccionadas:
        ws_resumen.append(["(sin cuentas)", "No hay cuentas para el filtro aplicado", "", 0, 0, 0, 0])
        _autosize_columns(ws_resumen)
        filename = "libro_mayor_consolidado_sin_datos.xlsx"
        return _xlsx_response(wb, filename)

    used_titles = {"Resumen"}
    for c in seleccionadas:
        codigo = str(getattr(c, "codigo", "") or "").strip()
        nombre = str(getattr(c, "nombre", "") or "").strip()
        tipo_c = str(getattr(c, "tipo", "") or "").strip()
        data = crud_asientos.obtener_libro_mayor_cuenta(
            db,
            codigo_cuenta=codigo,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
        )

        base_title = _safe_sheet_title(f"{codigo}-{nombre[:16]}", fallback=codigo or "Cuenta")
        title = base_title
        suffix = 1
        while title in used_titles:
            suffix += 1
            title = _safe_sheet_title(f"{base_title[:27]}_{suffix}", fallback=f"Cuenta_{suffix}")
        used_titles.add(title)

        ws = wb.create_sheet(title=title)
        _format_libro_mayor_excel(ws, data=data)

        ws_resumen.append(
            [
                codigo,
                nombre,
                tipo_c,
                len(data.get("rows", [])),
                float(data.get("total_debe") or 0),
                float(data.get("total_haber") or 0),
                float(data.get("saldo_final") or 0),
            ]
        )

    for r in ws_resumen.iter_rows(min_row=2, min_col=5, max_col=7):
        for cell in r:
            cell.number_format = '#,##0.00;[Red]-#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")
    _autosize_columns(ws_resumen)

    filename = (
        f"libro_mayor_consolidado_{fecha_desde or 'inicio'}_{fecha_hasta or 'hoy'}"
        f"{'_' + tipo_filter.lower() if tipo_filter else ''}.xlsx"
    )
    return _xlsx_response(wb, filename)


@router.get("/contabilidad/balance-general", response_class=HTMLResponse, name="contabilidad_balance_general")
def contabilidad_balance_general(
    request: Request,
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    resumen = crud_asientos.obtener_balance_general(db, fecha_hasta=fecha_hasta)
    ta: Decimal = resumen["total_activos"]
    tp: Decimal = resumen["total_pasivos"]
    tpat: Decimal = resumen["total_patrimonio"]
    pasivo_mas_patrimonio = tp + tpat
    diferencia_cuadre = ta - pasivo_mas_patrimonio
    fh = _parse_iso_date(fecha_hasta) or date.today()
    fd = _parse_iso_date(fecha_desde) or date(fh.year, 1, 1)
    if fd > fh:
        fd = date(fh.year, 1, 1)
    cuentas_movimiento = crud_plan.listar_cuentas_movimiento_activas(db)

    return templates.TemplateResponse(
        "finanzas/balance_general.html",
        {
            "request": request,
            "resumen": resumen,
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,
            "fecha_hoy": date.today().isoformat(),
            "bg_fecha_desde_default": fd.isoformat(),
            "bg_fecha_hasta_default": fh.isoformat(),
            "cuentas_movimiento": cuentas_movimiento,
            "tipos_cuenta": ["ACTIVO", "PASIVO", "PATRIMONIO", "INGRESO", "COSTO", "GASTO"],
            "pasivo_mas_patrimonio": pasivo_mas_patrimonio,
            "diferencia_cuadre": diferencia_cuadre,
            "active_menu": "contabilidad_balance_general",
        },
    )