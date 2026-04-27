# routes/ui/cobranza.py
# -*- coding: utf-8 -*-
from __future__ import annotations

import logging
from datetime import date, datetime
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, Form, HTTPException, Query, Request, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from core.paths import TEMPLATES_DIR
from core.public_errors import public_error_message
from core.rbac import guard_finanzas_consulta, guard_finanzas_mutacion
from crud.cobranza import cobranza as crud_cobranza
from crud.comunicaciones import email_log as crud_email_log
from db.session import get_db
from models import Caja, Cliente
from services.cobranza.pago_service import contabilizar_pago_cliente
from services.comunicaciones.email_service import enviar_recordatorio_cobranza

router = APIRouter(tags=["Cobranza"])
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
logger = logging.getLogger("evalua.cobranza.ui")


def _redirect(
    request: Request,
    route_name: str,
    *,
    msg: str | None = None,
    sev: str = "info",
    status_code: int = status.HTTP_303_SEE_OTHER,
    **path_params: Any,
) -> RedirectResponse:
    url = str(request.url_for(route_name, **path_params))
    if msg:
        sep = "&" if "?" in url else "?"
        url = f"{url}{sep}msg={msg}&sev={sev}"
    return RedirectResponse(url=url, status_code=status_code)


def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)


def _header_style(ws, row: int = 1) -> None:
    bold = Font(bold=True)
    for c in ws[row]:
        c.font = bold
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _en_horario_habil_chile(now: datetime | None = None) -> bool:
    now = now or datetime.now()
    # lunes=0 ... domingo=6
    if now.weekday() >= 5:
        return False
    return 9 <= now.hour < 18


@router.get("/cobranza", response_class=HTMLResponse, name="cobranza_resumen")
def cobranza_resumen(
    request: Request,
    msg: str | None = Query(None),
    sev: str = Query("info"),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    return templates.TemplateResponse(
        "cobranza/cobranza_resumen.html",
        {
            "request": request,
            "resumen_clientes": crud_cobranza.resumen_cobranza_por_cliente(db),
            "resumen_global": crud_cobranza.resumen_cobranza_general(db),
            "has_dashboard": True,
            "msg": msg,
            "sev": sev,
            "active_menu": "cobranza",
        },
    )


@router.get("/cobranza/dashboard", response_class=HTMLResponse, name="cobranza_dashboard")
def cobranza_dashboard(
    request: Request,
    msg: str | None = Query(None),
    sev: str = Query("info"),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    kpis = crud_cobranza.obtener_kpis_dashboard_cobranza(db)
    top_deudores = crud_cobranza.obtener_top_deudores_dashboard(db, limit=10)
    email_kpis = crud_cobranza.obtener_kpis_email_dashboard(db)
    ultimos_emails = crud_cobranza.obtener_ultimos_emails_dashboard(db, limit=12)
    aging = crud_cobranza.obtener_aging_saldos_dashboard(db)
    recuperacion = crud_cobranza.obtener_recuperacion_reciente_dashboard(db)
    proximos_venc = crud_cobranza.obtener_proximos_vencimientos_dashboard(db, limit=14)

    saldo_total = float(kpis.get("saldo_total_num") or 0)
    top3_saldo = sum(float(d.get("saldo_num") or 0) for d in (top_deudores or [])[:3])
    concentracion_top3_pct = (top3_saldo / saldo_total * 100.0) if saldo_total > 0 else 0.0

    return templates.TemplateResponse(
        "admin/cobranza_dashboard.html",
        {
            "request": request,
            "msg": msg,
            "sev": sev,
            "active_menu": "cobranza_dashboard",
            **kpis,
            **email_kpis,
            **aging,
            **recuperacion,
            "top_deudores": top_deudores,
            "ultimos_emails": ultimos_emails,
            "proximos_vencimientos": proximos_venc,
            "concentracion_top3_pct": concentracion_top3_pct,
            "concentracion_top3_fmt": f"{concentracion_top3_pct:.1f}".replace(".", ","),
            "concentracion_top3_monto": top3_saldo,
        },
    )


@router.get("/cobranza/export/cuentas.xlsx", name="cobranza_export_cuentas_excel")
def cobranza_export_cuentas_excel(
    request: Request,
    cliente_id: int | None = Query(None),
    solo_con_saldo: bool | None = Query(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    rows = crud_cobranza.export_cuentas_por_cobrar(
        db,
        cliente_id=cliente_id,
        solo_con_saldo=solo_con_saldo,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "CuentasPorCobrar"

    headers = [
        "CXC_ID",
        "CLIENTE_ID",
        "CLIENTE",
        "FECHA_EMISION",
        "FECHA_VENCIMIENTO",
        "ESTADO",
        "MONTO_ORIGINAL",
        "SALDO_PENDIENTE",
        "OBSERVACION",
    ]
    ws.append(headers)
    _header_style(ws)

    for r in rows:
        ws.append(
            [
                r.get("cxc_id"),
                r.get("cliente_id"),
                r.get("cliente"),
                str(r.get("fecha_emision") or ""),
                str(r.get("fecha_vencimiento") or ""),
                r.get("estado"),
                float(r.get("monto_original") or 0),
                float(r.get("saldo_pendiente") or 0),
                r.get("observacion") or "",
            ]
        )

    _autosize(ws)
    return _xlsx_response(wb, "cuentas_por_cobrar.xlsx")


@router.get("/cobranza/export/pagos.xlsx", name="cobranza_export_pagos_excel")
def cobranza_export_pagos_excel(
    request: Request,
    cliente_id: int | None = Query(None),
    cxc_id: int | None = Query(None),
    desde: date | None = Query(None),
    hasta: date | None = Query(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    rows = crud_cobranza.export_pagos_clientes(
        db,
        cliente_id=cliente_id,
        cxc_id=cxc_id,
        desde=desde,
        hasta=hasta,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos"

    headers = [
        "PAGO_ID",
        "FECHA_PAGO",
        "MONTO_PAGO",
        "FORMA_PAGO",
        "REFERENCIA",
        "OBSERVACION",
        "CAJA_ID",
        "CXC_ID",
        "FECHA_VENCIMIENTO_CXC",
        "ESTADO_CXC",
        "CLIENTE_ID",
        "CLIENTE",
    ]
    ws.append(headers)
    _header_style(ws)

    for r in rows:
        ws.append(
            [
                r.get("pago_id"),
                str(r.get("fecha_pago") or ""),
                float(r.get("monto_pago") or 0),
                r.get("forma_pago"),
                r.get("referencia") or "",
                r.get("observacion") or "",
                r.get("caja_id"),
                r.get("cxc_id"),
                str(r.get("fecha_vencimiento") or ""),
                r.get("estado_cxc"),
                r.get("cliente_id"),
                r.get("cliente"),
            ]
        )

    _autosize(ws)
    return _xlsx_response(wb, "pagos_clientes.xlsx")


@router.get("/cobranza/cliente/{cliente_id}", response_class=HTMLResponse, name="cobranza_detalle_cliente")
def cobranza_detalle_cliente(
    request: Request,
    cliente_id: int,
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    cliente = db.get(Cliente, cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    cuentas = crud_cobranza.listar_cuentas_por_cobrar_por_cliente(
        db,
        cliente_id=cliente_id,
        solo_con_saldo=False,
        incluir_pagos=True,
    )

    return templates.TemplateResponse(
        "cobranza/cobranza_detalle_cliente.html",
        {
            "request": request,
            "cliente": cliente,
            "cuentas": cuentas,
            "active_menu": "cobranza",
        },
    )


@router.get("/cobranza/gestion/{cliente_id}", response_class=HTMLResponse, name="cobranza_gestion")
def cobranza_gestion(
    request: Request,
    cliente_id: int,
    msg: str | None = Query(None),
    sev: str = Query("info"),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    cliente = db.get(Cliente, cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    cuentas = crud_cobranza.listar_cuentas_por_cobrar_por_cliente(
        db,
        cliente_id=cliente_id,
        solo_con_saldo=True,
        incluir_pagos=False,
    )

    logs_email = crud_email_log.listar_logs_cliente(db, cliente_id=cliente_id, limit=15)

    return templates.TemplateResponse(
        "cobranza/cobranza_gestion.html",
        {
            "request": request,
            "cliente": cliente,
            "cuentas": cuentas,
            "logs_email": logs_email,
            "msg": msg,
            "sev": sev,
            "active_menu": "cobranza",
        },
    )


@router.post("/cobranza/gestion/{cliente_id}/recordatorio", name="cobranza_enviar_recordatorio")
def cobranza_enviar_recordatorio(
    request: Request,
    cliente_id: int,
    incluir_detalle: int | None = Form(None),
    comentarios: str | None = Form(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_mutacion(request)) is not None:
        return redir
    cliente = db.get(Cliente, cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    if not cliente.email:
        return _redirect(
            request,
            "cobranza_gestion",
            cliente_id=cliente_id,
            msg="El cliente no tiene correo registrado.",
            sev="warning",
        )

    if not _en_horario_habil_chile():
        return _redirect(
            request,
            "cobranza_gestion",
            cliente_id=cliente_id,
            msg="La cobranza automatizada/manual está permitida solo en horario hábil de lunes a viernes.",
            sev="warning",
        )

    if crud_email_log.ya_enviado_hoy_cliente(db, cliente_id=cliente_id):
        return _redirect(
            request,
            "cobranza_gestion",
            cliente_id=cliente_id,
            msg="Ya se envió un recordatorio hoy a este cliente.",
            sev="warning",
        )

    cuentas = crud_cobranza.listar_cuentas_por_cobrar_por_cliente(
        db,
        cliente_id=cliente_id,
        solo_con_saldo=True,
        incluir_pagos=False,
    )

    if not cuentas:
        return _redirect(
            request,
            "cobranza_gestion",
            cliente_id=cliente_id,
            msg="El cliente no tiene documentos con saldo pendiente.",
            sev="warning",
        )

    try:
        enviar_recordatorio_cobranza(
            db=db,
            cliente=cliente,
            cuentas=cuentas,
            incluir_detalle=bool(incluir_detalle),
            comentarios=comentarios,
        )
        return _redirect(
            request,
            "cobranza_gestion",
            cliente_id=cliente_id,
            msg="Recordatorio enviado correctamente.",
            sev="success",
        )
    except Exception as exc:
        logger.exception("Enviar recordatorio cobranza cliente_id=%s", cliente_id)
        return _redirect(
            request,
            "cobranza_gestion",
            cliente_id=cliente_id,
            msg=public_error_message(exc, default="No fue posible enviar el recordatorio."),
            sev="danger",
        )


@router.post("/cobranza/automatizada/enviar", name="cobranza_automatizada_enviar")
def cobranza_automatizada_enviar(
    request: Request,
    dias_vencido: int = Form(0),
    incluir_detalle: int | None = Form(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_mutacion(request)) is not None:
        return redir
    if not _en_horario_habil_chile():
        return _redirect(
            request,
            "cobranza_resumen",
            msg="La cobranza automatizada solo puede ejecutarse en horario hábil de lunes a viernes.",
            sev="warning",
        )

    resumen_clientes = crud_cobranza.resumen_cobranza_por_cliente(db)
    enviados = 0
    omitidos = 0
    errores = 0

    for row in resumen_clientes:
        cliente_id = getattr(row, "cliente_id", None)
        if not cliente_id:
            omitidos += 1
            continue

        cliente = db.get(Cliente, cliente_id)
        if not cliente or not cliente.email:
            omitidos += 1
            continue

        if crud_email_log.ya_enviado_hoy_cliente(db, cliente_id=cliente_id):
            omitidos += 1
            continue

        cuentas = crud_cobranza.listar_cuentas_por_cobrar_por_cliente(
            db,
            cliente_id=cliente_id,
            solo_con_saldo=True,
            incluir_pagos=False,
        )

        cuentas_filtradas = []
        for c in cuentas:
            dias = getattr(c, "dias_vencido", None)
            if dias is None:
                fecha_vencimiento = getattr(c, "fecha_vencimiento", None)
                if fecha_vencimiento:
                    try:
                        dias = max((date.today() - fecha_vencimiento).days, 0)
                    except Exception:
                        dias = 0
                else:
                    dias = 0

            if int(dias or 0) >= int(dias_vencido or 0):
                cuentas_filtradas.append(c)

        if not cuentas_filtradas:
            omitidos += 1
            continue

        try:
            enviar_recordatorio_cobranza(
                db=db,
                cliente=cliente,
                cuentas=cuentas_filtradas,
                incluir_detalle=bool(incluir_detalle or 1),
            )
            enviados += 1
        except Exception:
            errores += 1

    sev = "success" if errores == 0 else "warning"
    msg = f"Proceso finalizado. Enviados={enviados}, omitidos={omitidos}, errores={errores}"

    return _redirect(
        request,
        "cobranza_resumen",
        msg=msg,
        sev=sev,
    )


@router.get("/cobranza/cuenta/{cxc_id}/pago", response_class=HTMLResponse, name="cobranza_pago_form")
def cobranza_pago_form(
    request: Request,
    cxc_id: int,
    msg: str | None = Query(None),
    sev: str = Query("info"),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_consulta(request)) is not None:
        return redir
    cxc = crud_cobranza.get_cuenta_por_cobrar(db, cxc_id)
    if not cxc:
        raise HTTPException(status_code=404, detail="Cuenta por cobrar no encontrada")

    cliente = db.get(Cliente, cxc.cliente_id) if cxc.cliente_id else None
    pagos = crud_cobranza.listar_pagos_por_cuenta(db, cxc_id=cxc_id)
    cajas = db.query(Caja).filter(Caja.activa.is_(True)).order_by(Caja.nombre.asc()).all()

    return templates.TemplateResponse(
        "cobranza/cobranza_pago_form.html",
        {
            "request": request,
            "cxc": cxc,
            "cliente": cliente,
            "pagos": pagos,
            "cajas": cajas,
            "hoy": date.today().isoformat(),
            "error": msg if sev in ("warning", "danger") else None,
            "active_menu": "cobranza",
        },
    )


@router.post("/cobranza/cuenta/{cxc_id}/pago", name="cobranza_registrar_pago")
def cobranza_registrar_pago(
    request: Request,
    cxc_id: int,
    fecha_pago: date = Form(...),
    monto_pago: float = Form(...),
    forma_pago: str = Form(...),
    caja_id: int | None = Form(None),
    referencia: str | None = Form(None),
    observacion: str | None = Form(None),
    db: Session = Depends(get_db),
):
    if (redir := guard_finanzas_mutacion(request)) is not None:
        return redir
    cxc = crud_cobranza.get_cuenta_por_cobrar(db, cxc_id)
    if not cxc:
        raise HTTPException(status_code=404, detail="Cuenta por cobrar no encontrada")

    try:
        pago = crud_cobranza.crear_pago(
            db=db,
            cxc=cxc,
            fecha_pago=fecha_pago,
            monto_pago=monto_pago,
            forma_pago=forma_pago,
            caja_id=caja_id,
            referencia=referencia,
            observacion=observacion,
        )
        auth = getattr(request.state, "auth_user", None)
        actor = None
        if isinstance(auth, dict):
            actor = str(auth.get("email") or auth.get("username") or auth.get("sub") or "")
        contabilizar_pago_cliente(db, pago_id=int(pago.id), usuario=actor, actualizar_cxc=False)
        return _redirect(
            request,
            "cobranza_detalle_cliente",
            cliente_id=cxc.cliente_id,
            msg="Pago registrado",
            sev="success",
        )
    except ValueError as exc:
        return _redirect(
            request,
            "cobranza_pago_form",
            cxc_id=cxc_id,
            msg=public_error_message(exc),
            sev="warning",
        )